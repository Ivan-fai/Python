import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt
#функція умовного форматування
def fill_cells(file, df):
    wb = load_workbook(file)
    ws = wb['Найкращі_Клієнти']
    #Визначаємо середній баланс всіх клієнтів
    avg_balance = df['balance'].mean()
    print(f"Середній баланс всіх клієнтів: {avg_balance:.2f}")
    lower = avg_balance - 100
    upper = avg_balance + 100
    #Визначаємо кольори
    Red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Червоний
    Yellow = PatternFill(start_color='FFFF00', end_color='FFFF00',fill_type='solid')  # Жовтий
    Green = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Зелений
    balance_ind = 4
    # Перебираємо всі рядки, починаючи з другого
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=balance_ind, max_col=balance_ind):
        cell = row[0]
        cell_value = cell.value
        if isinstance(cell_value, (int, float)):
            if cell_value < lower:
                cell.fill = Red
            elif lower <= cell_value <= upper:
                cell.fill = Yellow
            else:
                cell.fill = Green
    wb.save(file)
#головна функція програми
def  main():
    #побудова діаграми
    def plot_pie_chart(data, title):
        labels = data.index.tolist()
        sizes = data.values.tolist()
        plt.figure(figsize=(8, 8))
        plt.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
        plt.title(title, fontsize=14)
        plt.axis('equal')
        plt.show()
    #фільтрація даних
    def filtered(param):
        DF = df[df["y"] == param]
        return DF

    #початок головної функції
    csv_file= 'bank.csv'
    data = []
    try:
        df =  pd.read_csv(csv_file,sep = ';')
    except FileNotFoundError:
        print(f"Помилка: Файл не знайдено за шляхом {csv_file}")
        exit()
    print("Перевірка правильності зчитання датафрейму\n")
    print(df.head())

    print("\n--- Середні показники ---")
    #Середній вік
    avg_age_all = df['age'].mean()
    print(f"Середній вік : {avg_age_all:.2f}")
    #Середній баланс
    avg_balance_all = df['balance'].mean()
    print(f"Середній баланс: {avg_balance_all:.2f}")
    #Середня тривалість дзвінка
    avg_duration = df['duration'].mean()
    print(f"Середня тривалість дзвінка : {avg_duration:.2f} сек.")
    #Середня кількість контактів
    avg_campaign = df['campaign'].mean()
    print(f"Середня кількість контактів : {avg_campaign:.0f}")

    #занесення середніх значень до ексель аркуша
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    ws = wb.create_sheet(title="Середні значення")
    ws['B1'] = 'Середні показники'
    ws['A2'] = 'Середній вік'
    ws['A3'] = 'Середній баланс'
    ws['A4'] = 'Середня тривалість дзвінка'
    ws['A5'] = 'Середня кількість контактів'
    ws['B2'] = avg_age_all
    ws['B3'] = avg_balance_all
    ws['B4'] = avg_duration
    ws['B5'] = avg_campaign
    wb.save('bank_result.xlsx')

    # знаходження найкращих показників
    DF = filtered("yes")
    top_best_dur = DF.sort_values(by='duration', ascending=False)
    #знаходження найгірших показників
    DF_fail = filtered("no")
    top_failed_camp = DF_fail.sort_values(by='campaign', ascending=False)
    # Стовпці для ексль файлу
    best_leads = top_best_dur[['age', 'job', 'marital', 'balance', 'duration', 'campaign']]
    worst_leads = top_failed_camp[['age', 'job', 'marital', 'balance', 'duration', 'campaign']]

    best_100 = best_leads.head(100)
    worst_100 = worst_leads.head(100)
    excel_file = 'bank_result.xlsx'
    # Використовуємо pd.ExcelWriter для запису на різні листи
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode = 'a') as writer:
        best_100.to_excel(writer, sheet_name='Найкращі_Клієнти', index=False)#запис 100 найкращих клієнтів
        worst_100.to_excel(writer, sheet_name='Найгірші_Клієнти', index=False)#запис 100 найгірших клієнтів
        print(f"\nДані успішно записано у файл '{excel_file}' на два листи.")
    #вибір для користувача чи застосовувати умовне форматування кольором
    print("Застосовано умовне форматування до стовпця balance")
    fill_cells('bank_result.xlsx', df)
    #побудова грфіка функції на вибір за користувачем
    print("Побудова графіка функцї")
    choice = int(input("Оберіть число для побудови діаграми: 1 - За 3 найпопулярнішими професіями або 2 - За 3 видами навчання:"))
    while True:
        if choice == 1:
            f = DF['job'].value_counts().head(3)
            print("Побудовано діаграму '3 найбільш популярні професії'")
            plot_pie_chart(f, "3 найбільш популярні професії")
            plt.savefig('pie_chart1.png')
            break
        if choice == 2:
            f1 = DF['education'].value_counts().head(3)
            print("Побудовано діаграму '3 найбільш розповсюджені рівні навчання'")
            plot_pie_chart(f1, "3 найбільш розповсюджені рівні навчання")
            plt.savefig('pie_chart2.png')
            break
        else: choice = int(input("Введено некоректне значення, будь ласка оберіть 1 або 2: "))
main()